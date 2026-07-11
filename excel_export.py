import io
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import models

# Premium Color Palette
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
HEADER_FONT = Font(name="Cordia New", size=14, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Cordia New", size=18, bold=True, color="1E1B4B") # Very dark indigo
SUBTITLE_FONT = Font(name="Cordia New", size=12, italic=True, color="475569")
SECTION_FONT = Font(name="Cordia New", size=14, bold=True, color="312E81")
LABEL_FONT = Font(name="Cordia New", size=12, bold=True, color="1E293B")
VALUE_FONT = Font(name="Cordia New", size=12, color="334155")
SUMMARY_FONT = Font(name="Cordia New", size=13, bold=True, color="000000")

# Borders
THIN_SIDE = Side(border_style="thin", color="CBD5E1")
BORDER_ALL = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
BORDER_TOP_THIN_BOTTOM_DOUBLE = Border(
    top=Side(border_style="thin", color="94A3B8"),
    bottom=Side(border_style="double", color="475569")
)

def format_thai_date(d):
    if not d:
        return "-"
    if isinstance(d, (date, datetime)):
        months = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
        return f"{d.day} {months[d.month-1]} {d.year + 543}"
    return str(d)

def autofit_columns(ws, max_len_adjust=4):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            # Handle Thai characters length properly in Excel column sizing
            val_len = len(val)
            if val_len > max_len:
                max_len = val_len
        ws.column_dimensions[col_letter].width = max(max_len + max_len_adjust, 12)

def export_projects_to_excel(projects):
    wb = Workbook()
    ws = wb.active
    ws.title = "ภาพรวมโครงการสัญญา"
    
    # Title Block
    ws.append([])
    ws["B2"] = "รายงานสรุปภาพรวมโครงการสัญญาและงบประมาณ"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"วันที่ออกรายงาน: {format_thai_date(date.today())}"
    ws["B3"].font = SUBTITLE_FONT
    ws.append([])
    ws.append([])
    
    headers = [
        "ลำดับ", "ชื่อโครงการ", "เจ้าของโครงการ / หน่วยงาน", "บริษัทที่รับผิดชอบ", 
        "ประเภทงาน", "ปีงบประมาณ", "งบประมาณรวม (บาท)", "ระยะเวลาดำเนินงานสัญญา (วัน)", 
        "วันเริ่มต้นสัญญา", "วันสิ้นสุดสัญญา", "สถานะ"
    ]
    
    ws.append(headers)
    header_row = ws.max_row
    
    # Style Header Row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
    
    ws.row_dimensions[header_row].height = 28
    
    start_row = header_row + 1
    for idx, p in enumerate(projects, 1):
        row_data = [
            idx,
            p.name,
            p.owner,
            p.contractor or "-",
            p.job_type or "-",
            p.fiscal_year or "-",
            p.budget,
            p.contract_duration_days or "-",
            format_thai_date(p.start_date),
            format_thai_date(p.end_date),
            p.status
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        
        # Styles for values
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = VALUE_FONT
            cell.border = BORDER_ALL
            
            # Alignments & Number formats
            if col_idx in [1, 6, 8, 9, 10, 11]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 7: # Budget
                cell.number_format = '฿#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
    end_row = ws.max_row
    
    # Summary Row
    ws.append([None])
    summary_row = ws.max_row
    ws.cell(row=summary_row, column=1, value="รวมทั้งสิ้น").font = SUMMARY_FONT
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="center")
    
    # Total Projects count
    ws.cell(row=summary_row, column=2, value=f"{len(projects)} โครงการ").font = SUMMARY_FONT
    
    # Sum Formula for Budget (Column G is index 7)
    budget_cell = ws.cell(row=summary_row, column=7)
    budget_cell.value = f"=SUM(G{start_row}:G{end_row})"
    budget_cell.font = SUMMARY_FONT
    budget_cell.number_format = '฿#,##0.00'
    budget_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # Apply border to summary row
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=summary_row, column=col_idx).border = BORDER_TOP_THIN_BOTTOM_DOUBLE
        
    autofit_columns(ws)
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def export_purchase_orders_to_excel(pos):
    wb = Workbook()
    ws = wb.active
    ws.title = "รายการใบสั่งซื้อ PO"
    
    # Title Block
    ws.append([])
    ws["B2"] = "รายงานสรุปรายการใบสั่งซื้อวัสดุ (Purchase Orders)"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"วันที่ออกรายงาน: {format_thai_date(date.today())}"
    ws["B3"].font = SUBTITLE_FONT
    ws.append([])
    ws.append([])
    
    headers = [
        "ลำดับ", "เลขที่ PO", "ชื่อโครงการสัญญา", "เจ้าของโครงการ / หน่วยงาน", 
        "งบประมาณ PO (บาท)", "กำหนดส่งมอบ", "ผู้รับผิดชอบ (จัดซื้อ)", "สถานะส่งมอบ"
    ]
    
    ws.append(headers)
    header_row = ws.max_row
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL
        
    ws.row_dimensions[header_row].height = 28
    
    start_row = header_row + 1
    for idx, po in enumerate(pos, 1):
        row_data = [
            idx,
            po.po_number,
            po.project.name if po.project else "-",
            po.project.owner if po.project else "-",
            po.budget,
            format_thai_date(po.due_date),
            po.contractor,
            po.delivery_status
        ]
        ws.append(row_data)
        curr_row = ws.max_row
        ws.row_dimensions[curr_row].height = 20
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=curr_row, column=col_idx)
            cell.font = VALUE_FONT
            cell.border = BORDER_ALL
            
            if col_idx in [1, 2, 6, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 5: # PO Budget
                cell.number_format = '฿#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
    end_row = ws.max_row
    
    # Summary Row
    ws.append([None])
    summary_row = ws.max_row
    ws.cell(row=summary_row, column=1, value="รวมทั้งสิ้น").font = SUMMARY_FONT
    ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=summary_row, column=2, value=f"{len(pos)} รายการ PO").font = SUMMARY_FONT
    
    budget_cell = ws.cell(row=summary_row, column=5)
    budget_cell.value = f"=SUM(E{start_row}:E{end_row})"
    budget_cell.font = SUMMARY_FONT
    budget_cell.number_format = '฿#,##0.00'
    budget_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=summary_row, column=col_idx).border = BORDER_TOP_THIN_BOTTOM_DOUBLE
        
    autofit_columns(ws)
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def export_project_detail_to_excel(project):
    wb = Workbook()
    
    # Sheet 1: Project Metadata
    ws1 = wb.active
    ws1.title = "ข้อมูลโครงการ"
    
    ws1.append([])
    ws1["B2"] = "รายงานรายละเอียดโครงการสัญญาและคู่สัญญา"
    ws1["B2"].font = TITLE_FONT
    ws1["B3"] = f"พิมพ์ ณ วันที่: {format_thai_date(date.today())}"
    ws1["B3"].font = SUBTITLE_FONT
    ws1.append([])
    ws1.append([])
    
    # Detail Key-Values
    metadata = [
        ("ชื่อโครงการสัญญา", project.name),
        ("เจ้าของโครงการ / หน่วยงาน", project.owner),
        ("เลขที่สัญญา", project.contract_number or "-"),
        ("บริษัทผู้รับผิดชอบ", project.contractor or "-"),
        ("ประเภทงาน", project.job_type or "-"),
        ("ปีงบประมาณ", project.fiscal_year or "-"),
        ("งบประมาณรวมโครงการ", project.budget),
        ("ระยะเวลาดำเนินงานสัญญา (วัน)", project.contract_duration_days or "-"),
        ("วันที่เซ็นสัญญา", format_thai_date(project.contract_signing_date)),
        ("วันที่เริ่มต้นสัญญา", format_thai_date(project.start_date)),
        ("วันที่สิ้นสุดสัญญา", format_thai_date(project.end_date)),
        ("วันที่สั่งเข้างานจริง", format_thai_date(project.work_order_date)),
        ("สถานะปัจจุบัน", project.status),
        ("การส่งมอบคู่ฉบับสัญญา", project.counterpart_status),
        ("การโอนสิทธิ์", project.right_assignment),
        ("สัดส่วนการโอนสิทธิ์ (%)", f"{project.right_assignment_percentage}%" if project.right_assignment_percentage else "-"),
        ("วงเงินหลักประกันสัญญา", project.guarantee_amount),
        ("รูปแบบหลักประกัน", project.guarantee_payment_type),
        ("ธนาคารผู้ค้ำประกัน (LG)", project.guarantee_bank or "-"),
        ("วันหมดอายุหนังสือค้ำประกัน", format_thai_date(project.guarantee_expiry_date)),
        ("สถานะใบเสร็จหลักประกัน", project.guarantee_receipt_status),
        ("เลขที่ใบเสร็จหลักประกัน", project.guarantee_receipt_number or "-")
    ]
    
    curr_row = 5
    for key, val in metadata:
        ws1.cell(row=curr_row, column=2, value=key).font = LABEL_FONT
        ws1.cell(row=curr_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws1.cell(row=curr_row, column=2).border = BORDER_ALL
        
        val_cell = ws1.cell(row=curr_row, column=3, value=val)
        val_cell.font = VALUE_FONT
        val_cell.border = BORDER_ALL
        
        if key in ["งบประมาณรวมโครงการ", "วงเงินหลักประกันสัญญา"]:
            val_cell.number_format = '฿#,##0.00'
            val_cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            val_cell.alignment = Alignment(horizontal="left", vertical="center")
            
        ws1.row_dimensions[curr_row].height = 20
        curr_row += 1
        
    autofit_columns(ws1, max_len_adjust=6)
    
    # Sheet 2: Deliverables
    ws2 = wb.create_sheet(title="งวดงานส่งมอบ")
    ws2.append([])
    ws2["B2"] = f"รายการงวดงานส่งมอบ: {project.name}"
    ws2["B2"].font = TITLE_FONT
    ws2.append([])
    ws2.append([])
    
    deliv_headers = [
        "ลำดับ", "รายการส่งมอบ / วัสดุ", "งวดงาน", "งบประมาณงวดงาน (บาท)", 
        "เลขที่ใบส่งของ", "เลขที่ส่งภายใน", "เลขที่ส่งภายนอก", "กำหนดวันส่งมอบ", "สถานะการส่งมอบ"
    ]
    
    ws2.append(deliv_headers)
    h_row = ws2.max_row
    for col_idx in range(1, len(deliv_headers) + 1):
        c = ws2.cell(row=h_row, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL
        
    ws2.row_dimensions[h_row].height = 26
    
    deliv_start = ws2.max_row + 1
    for idx, d in enumerate(project.deliverables, 1):
        ws2.append([
            idx,
            d.name,
            d.milestone or "-",
            d.budget or 0,
            d.delivery_no or "-",
            d.internal_delivery_no or "-",
            d.external_delivery_no or "-",
            format_thai_date(d.due_date),
            d.status
        ])
        r = ws2.max_row
        ws2.row_dimensions[r].height = 20
        
        for col_idx in range(1, len(deliv_headers) + 1):
            cell = ws2.cell(row=r, column=col_idx)
            cell.font = VALUE_FONT
            cell.border = BORDER_ALL
            if col_idx in [1, 8, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 4:
                cell.number_format = '฿#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
    deliv_end = ws2.max_row
    
    if len(project.deliverables) > 0:
        ws2.append([None])
        sum_row = ws2.max_row
        ws2.cell(row=sum_row, column=1, value="รวมงบประมาณ").font = SUMMARY_FONT
        ws2.cell(row=sum_row, column=1).alignment = Alignment(horizontal="center")
        
        sum_cell = ws2.cell(row=sum_row, column=4)
        sum_cell.value = f"=SUM(D{deliv_start}:D{deliv_end})"
        sum_cell.font = SUMMARY_FONT
        sum_cell.number_format = '฿#,##0.00'
        sum_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        for col_idx in range(1, len(deliv_headers) + 1):
            ws2.cell(row=sum_row, column=col_idx).border = BORDER_TOP_THIN_BOTTOM_DOUBLE
            
    autofit_columns(ws2)
    
    # Sheet 3: Purchase Orders
    ws3 = wb.create_sheet(title="ใบสั่งซื้อวัสดุ PO")
    ws3.append([])
    ws3["B2"] = f"รายการใบสั่งซื้อวัสดุภายใต้โครงการ: {project.name}"
    ws3["B2"].font = TITLE_FONT
    ws3.append([])
    ws3.append([])
    
    po_headers = [
        "ลำดับ", "เลขที่ PO", "ประเภทวัสดุ / รายการจัดซื้อ", "งบประมาณ PO (บาท)", 
        "กำหนดส่งมอบ", "ผู้รับผิดชอบจัดซื้อ", "สถานะการส่งของ", "เลขที่ใบส่งของ", "วันที่ส่งของสำเร็จ"
    ]
    
    ws3.append(po_headers)
    h_row_3 = ws3.max_row
    for col_idx in range(1, len(po_headers) + 1):
        c = ws3.cell(row=h_row_3, column=col_idx)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER_ALL
        
    ws3.row_dimensions[h_row_3].height = 26
    
    po_start = ws3.max_row + 1
    for idx, p in enumerate(project.purchase_orders, 1):
        ws3.append([
            idx,
            p.po_number,
            p.material_type,
            p.budget,
            format_thai_date(p.due_date),
            p.contractor,
            p.delivery_status,
            p.delivery_no or "-",
            format_thai_date(p.delivery_date)
        ])
        r = ws3.max_row
        ws3.row_dimensions[r].height = 20
        
        for col_idx in range(1, len(po_headers) + 1):
            cell = ws3.cell(row=r, column=col_idx)
            cell.font = VALUE_FONT
            cell.border = BORDER_ALL
            if col_idx in [1, 2, 5, 7, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if col_idx == 4:
                cell.number_format = '฿#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
    po_end = ws3.max_row
    
    if len(project.purchase_orders) > 0:
        ws3.append([None])
        sum_row = ws3.max_row
        ws3.cell(row=sum_row, column=1, value="รวมงบประมาณ PO").font = SUMMARY_FONT
        ws3.cell(row=sum_row, column=1).alignment = Alignment(horizontal="center")
        
        sum_cell = ws3.cell(row=sum_row, column=4)
        sum_cell.value = f"=SUM(D{po_start}:D{po_end})"
        sum_cell.font = SUMMARY_FONT
        sum_cell.number_format = '฿#,##0.00'
        sum_cell.alignment = Alignment(horizontal="right", vertical="center")
        
        for col_idx in range(1, len(po_headers) + 1):
            ws3.cell(row=sum_row, column=col_idx).border = BORDER_TOP_THIN_BOTTOM_DOUBLE
            
    autofit_columns(ws3)
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out

def export_po_detail_to_excel(po):
    wb = Workbook()
    ws = wb.active
    ws.title = "รายละเอียดใบสั่งซื้อ PO"
    
    ws.append([])
    ws["B2"] = "รายงานรายละเอียดใบสั่งซื้อวัสดุ (Purchase Order)"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"พิมพ์ ณ วันที่: {format_thai_date(date.today())}"
    ws["B3"].font = SUBTITLE_FONT
    ws.append([])
    ws.append([])
    
    metadata = [
        ("เลขที่ใบสั่งซื้อ PO", po.po_number),
        ("โครงการสัญญาอ้างอิง", po.project.name if po.project else "-"),
        ("เจ้าของโครงการ / หน่วยงานผู้ว่าจ้าง", po.project.owner if po.project else "-"),
        ("ประเภทวัสดุ / รายการจัดซื้อ", po.material_type),
        ("งบประมาณใบสั่งซื้อ PO", po.budget),
        ("ผู้รับผิดชอบจัดซื้อ", po.contractor),
        ("กำหนดเวลาส่งมอบสินค้า", format_thai_date(po.due_date)),
        ("สถานะการจัดส่ง", po.delivery_status),
        ("เลขที่ใบส่งของ / ส่งมอบของ", po.delivery_no or "-"),
        ("วันที่ส่งมอบสินค้าจริง", format_thai_date(po.delivery_date)),
        ("ชื่อไฟล์แนบใบสั่งซื้อ PO", po.po_file_filename or "-"),
        ("ชื่อไฟล์แนบใบเสนอราคา (Quotation)", po.quotation_file_filename or "-"),
        ("ชื่อไฟล์แนบใบส่งของ (Delivery Invoice)", po.delivery_file_filename or "-")
    ]
    
    curr_row = 5
    for key, val in metadata:
        ws.cell(row=curr_row, column=2, value=key).font = LABEL_FONT
        ws.cell(row=curr_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=curr_row, column=2).border = BORDER_ALL
        
        val_cell = ws.cell(row=curr_row, column=3, value=val)
        val_cell.font = VALUE_FONT
        val_cell.border = BORDER_ALL
        
        if key == "งบประมาณใบสั่งซื้อ PO":
            val_cell.number_format = '฿#,##0.00'
            val_cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            val_cell.alignment = Alignment(horizontal="left", vertical="center")
            
        ws.row_dimensions[curr_row].height = 20
        curr_row += 1
        
    autofit_columns(ws, max_len_adjust=6)
    
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
